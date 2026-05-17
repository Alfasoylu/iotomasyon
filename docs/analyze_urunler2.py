# -*- coding: utf-8 -*-
import sys
import io

# Force UTF-8 output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
import pandas as pd

filepath = r"C:\dev\iotomasyoncom\iotomasyon\docs\urunler.xlsx"

print("=" * 80)
print("STEP 1: openpyxl with data_only=False (formulas)")
print("=" * 80)

try:
    wb_formulas = openpyxl.load_workbook(filepath, data_only=False)
    print(f"Sheet names: {wb_formulas.sheetnames}")
    print()

    # Named ranges
    print("Named ranges:")
    if wb_formulas.defined_names:
        for name, defn in wb_formulas.defined_names.items():
            print(f"  {name}: {defn.attr_text}")
    else:
        print("  (none)")
    print()

    for sheet_name in wb_formulas.sheetnames:
        ws = wb_formulas[sheet_name]
        print(f"--- Sheet: '{sheet_name}' (dims: {ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column}) ---")
        print("First 5 rows (formulas):")
        for i, row in enumerate(ws.iter_rows(values_only=False)):
            if i >= 5:
                break
            row_data = []
            for cell in row:
                val = cell.value
                row_data.append(f"{cell.coordinate}={repr(val)}")
            print("  " + " | ".join(row_data))
        print()
except Exception as e:
    print(f"Error with data_only=False: {e}")

print("=" * 80)
print("STEP 2: openpyxl with data_only=True (computed values) - ALL SHEETS FULL DATA")
print("=" * 80)

try:
    wb_values = openpyxl.load_workbook(filepath, data_only=True)
    print(f"Sheet names: {wb_values.sheetnames}")
    print()

    for sheet_idx, sheet_name in enumerate(wb_values.sheetnames):
        ws = wb_values[sheet_name]
        print(f"=== Sheet [{sheet_idx+1}]: '{sheet_name}' ===")
        print(f"    Dimensions: {ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column}")
        max_rows = min(ws.max_row or 0, 50)
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True)):
            if all(v is None for v in row) and i > 2:
                continue
            print(f"  Row {i+1}: {list(row)}")
        print()
except Exception as e:
    print(f"Error with data_only=True: {e}")

print("=" * 80)
print("STEP 3: pandas pd.read_excel - all sheets")
print("=" * 80)

try:
    all_sheets = pd.read_excel(filepath, sheet_name=None, header=None, dtype=str)
    print(f"Sheets found by pandas: {list(all_sheets.keys())}")
    print()
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', 400)
    pd.set_option('display.max_colwidth', 120)
    pd.set_option('display.max_rows', 60)
    for sheet_name, df in all_sheets.items():
        print(f"=== pandas Sheet: '{sheet_name}' ===")
        print(f"    Shape: {df.shape}")
        print(df.head(50).to_string())
        print()
except Exception as e:
    print(f"Error with pandas: {e}")

print("=" * 80)
print("DONE")
print("=" * 80)
